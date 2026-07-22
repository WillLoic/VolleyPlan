"""def generer_excel_planning(planning_id, coach_id) -> BytesIO:
    
    - Vérifie que le planning appartient au coach
    - Construit le classeur openpyxl avec les 3 feuilles décrites ci-dessus
    - Applique un style basique cohérent avec la charte VolleyPlan
      (en-têtes fond rouge/charcoal, texte blanc, colonnes auto-ajustées)
    - Retourne le fichier en mémoire (BytesIO) prêt à être streamé
    """


from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models.planning import Planning
from ..models.presences import Absence
from ..services.planning import PlanningService
from ..services.bilan import BilanService, DOMAINES

# ── Styles réutilisables ────────────────────────────────────────────
HEADER_FILL   = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
RED_FILL      = PatternFill(start_color="D72638", end_color="D72638", fill_type="solid")
TITLE_FONT    = Font(bold=True, size=14, color="D72638")
BOLD_FONT     = Font(bold=True)
THIN_BORDER   = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)

STATUT_COLORS = {
    "Effectuée":     PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "Non effectuée": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "Planifiée":     PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
}


def _style_header_row(ws, row_idx, nb_cols):
    for col in range(1, nb_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _autofit_columns(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _get_statut_seance(seance) -> str:
    aujourd_hui = date.today()
    if seance.date_seance is None:
        return "Planifiée"
    if seance.date_seance > aujourd_hui:
        return "Planifiée"
    if seance.presences_prises:
        return "Effectuée"
    return "Non effectuée"


def generer_excel_planning(planning_id: int, coach_id: int) -> tuple[BytesIO | None, str | None]:
    """
    Génère un classeur Excel à 3 feuilles pour un planning donné.
    Retourne (buffer, None) en cas de succès, (None, message_erreur) sinon.
    """
    planning, error = PlanningService.get_by_id(planning_id, coach_id)
    if error:
        return None, error

    seances = planning.seances or []
    joueurs = planning.joueurs or []

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════
    # FEUILLE 1 — PLANNING DÉTAILLÉ
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Planning"

    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"VolleyPlan — {planning.titre}"
    ws1["A1"].font = TITLE_FONT
    ws1.row_dimensions[1].height = 24

    headers1 = ["Séance", "Date", "Heure", "Domaines", "Exercices", "Durée totale (min)"]
    for i, h in enumerate(headers1, start=1):
        ws1.cell(row=3, column=i, value=h)
    _style_header_row(ws1, 3, len(headers1))

    row = 4
    for s in seances:
        domaines_labels = [
            d["label"] for d in DOMAINES if d["id"] in (s.domaines or [])
        ]
        exercices_txt = "\n".join(
            f"• {e.nom} ({e.duree}min)" for e in (s.exercices or [])
        )
        date_txt = s.date_seance.strftime("%d/%m/%Y") if s.date_seance else "—"

        ws1.cell(row=row, column=1, value=f"{s.ordre + 1}. {s.titre}").font = BOLD_FONT
        ws1.cell(row=row, column=2, value=date_txt)
        ws1.cell(row=row, column=3, value=s.heure_debut or "—")
        ws1.cell(row=row, column=4, value=", ".join(domaines_labels) or "—")
        ws1.cell(row=row, column=5, value=exercices_txt or "—")
        ws1.cell(row=row, column=6, value=s.duree_totale())

        for col in range(1, 7):
            cell = ws1.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = LEFT
        row += 1

    _autofit_columns(ws1, {
        "A": 22, "B": 12, "C": 10, "D": 26, "E": 40, "F": 16
    })

    # ═══════════════════════════════════════════════════════════════
    # FEUILLE 2 — BILAN
    # ═══════════════════════════════════════════════════════════════
    bilan, bilan_error = BilanService.compute(coach_id, planning_id)

    ws2 = wb.create_sheet("Bilan")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = f"Bilan — {planning.titre}"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 24

    if bilan:
        # Résumé en haut
        ws2["A3"] = "Nombre de séances :"
        ws2["B3"] = bilan["nb_seances"]
        ws2["A4"] = "Volume total :"
        ws2["B4"] = f"{bilan['total_minutes']} min"
        ws2["A5"] = "Durée moyenne / séance :"
        ws2["B5"] = f"{bilan['avg_seance_minutes']} min"
        for r in (3, 4, 5):
            ws2.cell(row=r, column=1).font = BOLD_FONT

        # Tableau répartition par domaine
        headers2 = ["Domaine", "Volume (min)", "% du total"]
        header_row2 = 7
        for i, h in enumerate(headers2, start=1):
            ws2.cell(row=header_row2, column=i, value=h)
        _style_header_row(ws2, header_row2, len(headers2))

        row = header_row2 + 1
        for d in bilan["domain_stats"]:
            ws2.cell(row=row, column=1, value=d["label"])
            ws2.cell(row=row, column=2, value=d["minutes"])
            ws2.cell(row=row, column=3, value=f"{d['pct']}%")
            for col in range(1, 4):
                ws2.cell(row=row, column=col).border = THIN_BORDER
                ws2.cell(row=row, column=col).alignment = CENTER
            row += 1

        # Recommandations
        row += 1
        ws2.cell(row=row, column=1, value="Recommandations").font = BOLD_FONT
        row += 1
        for rec in bilan["recommandations"]:
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws2.cell(row=row, column=1, value=f"• {rec}").alignment = LEFT
            row += 1

    _autofit_columns(ws2, {"A": 22, "B": 16, "C": 14, "D": 14})

    # ═══════════════════════════════════════════════════════════════
    # FEUILLE 3 — PRÉSENCES
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Présences")

    nb_cols_fixed = 3  # Séance, Date, Statut
    nb_cols = nb_cols_fixed + len(joueurs)

    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(nb_cols, 1))
    ws3["A1"] = f"Présences — {planning.titre}"
    ws3["A1"].font = TITLE_FONT
    ws3.row_dimensions[1].height = 24

    headers3 = ["Séance", "Date", "Statut"] + [j.nom for j in joueurs]
    for i, h in enumerate(headers3, start=1):
        ws3.cell(row=3, column=i, value=h)
    _style_header_row(ws3, 3, len(headers3))

    # Pré-charge toutes les absences des séances de ce planning en une requête
    seance_ids = [s.id for s in seances]
    absences = (
        Absence.query.filter(Absence.seance_id.in_(seance_ids)).all()
        if seance_ids else []
    )
    absences_par_seance = {}
    for a in absences:
        absences_par_seance.setdefault(a.seance_id, set()).add(a.joueur_id)

    row = 4
    for s in seances:
        statut = _get_statut_seance(s)
        date_txt = s.date_seance.strftime("%d/%m/%Y") if s.date_seance else "—"

        ws3.cell(row=row, column=1, value=f"{s.ordre + 1}. {s.titre}").font = BOLD_FONT
        ws3.cell(row=row, column=2, value=date_txt).alignment = CENTER
        statut_cell = ws3.cell(row=row, column=3, value=statut)
        statut_cell.alignment = CENTER
        statut_cell.fill = STATUT_COLORS.get(statut, STATUT_COLORS["Planifiée"])

        joueurs_absents = absences_par_seance.get(s.id, set())
        for col_offset, j in enumerate(joueurs):
            col = nb_cols_fixed + 1 + col_offset
            if statut == "Planifiée":
                symbole = "—"
            elif j.id in joueurs_absents:
                symbole = "✗"
            else:
                symbole = "✓"
            cell = ws3.cell(row=row, column=col, value=symbole)
            cell.alignment = CENTER

        for col in range(1, nb_cols + 1):
            ws3.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    widths3 = {"A": 22, "B": 12, "C": 16}
    for i in range(len(joueurs)):
        widths3[get_column_letter(nb_cols_fixed + 1 + i)] = 14
    _autofit_columns(ws3, widths3)

    # ═══════════════════════════════════════════════════════════════
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, None